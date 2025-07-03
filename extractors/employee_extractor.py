import os
import json
import pandas as pd
from datetime import datetime
from verint_client import VerintClient
from extractors.group_extractor import extract_groups

def parse_employee_skills(skill_json):
    """
    Return a list of *active* skills.

    A skill is considered active when **end_date** / **endDate** is
    • missing, "", None, or "null", **OR**
    • on or after today's UTC date (compare just the YYYY‑MM‑DD part).

    The function tolerates both camelCase and snake_case keys.
    """
    today = datetime.utcnow().date()

    skill_assignments = skill_json.get("data", [])
    active_skills: list[dict] = []
    dropped = 0

    for item in skill_assignments:
        attr = item.get("attributes", {}) or {}
        skill_meta = (
            item.get("relationships", {})
                .get("skill", {})
                .get("data", {})
                .get("meta", {})
        )

        # Handle snake/camel case for dates
        end_raw = attr.get("end_date") or attr.get("endDate") or ""
        start_raw = attr.get("start_date") or attr.get("startDate") or ""

        # Decide if we keep this assignment
        keep = False
        if not end_raw or str(end_raw).lower() == "null":
            keep = True
        else:
            try:
                end_dt = datetime.strptime(end_raw[:10], "%Y-%m-%d").date()
                keep = end_dt >= today
            except Exception:
                # If parsing fails, assume it's still relevant
                keep = True

        if keep:
            active_skills.append({
                "name": skill_meta.get("name", "null"),
                "proficiency": attr.get("proficiency", "null"),
                "priority": attr.get("priority", "null"),
                "start_date": start_raw,
                "end_date": end_raw,
                "reserve_level": (
                    attr.get("reserveLevel")
                    or attr.get("reserve_level")
                    or "null"
                ),
            })
        else:
            dropped += 1

    # Debug summary
    if skill_assignments:
        print(f"    kept {len(active_skills)} / {len(skill_assignments)} skills "
              f"(dropped {dropped})")

    return active_skills if active_skills else ""

def parse_employee_udfs(udf_json):
    udf_entries = udf_json.get("data", [])
    parsed_udfs = []
    for item in udf_entries:
        attr = item.get("attributes", {})
        parsed_udfs.append({
            "name": attr.get("name", "null"),
            "value": attr.get("value", "null")
        })
    return parsed_udfs if parsed_udfs else ""

def extract_employees(employee_groups_map):
    client = VerintClient()
    response = client.verint_call("wfo/user-mgmt-api/v1/employees")

    employees = response.get("data", [])
    records = []
    employee_types = set()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs("json_dump", exist_ok=True)
    with open(f"json_dump/employee_response_{timestamp}.json", "w") as f:
        json.dump(response, f, indent=4)

    data_source_cache = {}

    for emp in employees:
        attr = emp.get("attributes", {}) or {}
        person = attr.get("person") or {}
        contact = person.get("contact") or {}
        address = person.get("address") or {}
        user = attr.get("user") or {}
        org_data = emp.get("relationships", {}).get("organization", {}).get("data") or {}
        org_meta = org_data.get("meta") or {}
        emp_type = attr.get("employeeType")
        if emp_type:
            employee_types.add(emp_type)

        employee_id = emp.get("id")
        organization_id = attr.get("organizationId")

        try:
            job_res = client.verint_call(f"wfo/user-mgmt-api/v1/employees/{employee_id}/jobTitle")
            job_data = job_res.get("data", {})
            job_id = job_data.get("id")
            job_name = job_data.get("attributes", {}).get("name")
            job_title_json = json.dumps({"id": job_id, "name": job_name}) if job_id and job_name else None
        except Exception as e:
            print(f"Job title fetch failed for Employee ID {employee_id} — skipping. Error: {e}")
            job_title_json = None

        workspace_logins = []
        try:
            workspace_res = client.verint_call(f"wfo/user-mgmt-api/v1/employees/{employee_id}/workspace")
            assets = workspace_res.get("data", {}).get("attributes", {}).get("assets", [])
            for asset in assets:
                ds_id = str(asset.get("dataSourceID"))
                login_name = asset.get("loginName") or "null"
                if ds_id not in data_source_cache:
                    ds_res = client.verint_call(f"api/em/v2/datasources/{ds_id}")
                    ds_name = ds_res.get("data", [{}])[0].get("attributes", {}).get("name", ds_id)
                    data_source_cache[ds_id] = ds_name
                workspace_logins.append(f"{data_source_cache[ds_id]} - {login_name}")
        except Exception as e:
            print(f"Workspace fetch failed for Employee ID {employee_id} — skipping. Error: {e}")

        # Updated preference handling
        pref_keys = ",".join([
            "UserTimezone", "UserLanguage", "UserDefaultPageRows",
            "UserRegionalFormat", "UserAccessibilityComplianceMode",
            "UserShowOrgListHierarchical", "UserRepeatHeaderInterval",
            "UserLoginScreenName"
        ])

        parsed_preferences = []

        try:
            prefs_res = client.verint_call(
                f"wfo/user-mgmt-api/v1/employees/{employee_id}/preferences?keys={pref_keys}"
            )
            for pref in prefs_res.get("data", []):
                key = pref.get("id")
                value = pref.get("attributes", {}).get("value")
                if value is not None and value != "null":
                    parsed_preferences.append({
                        "name": key,
                        "value": value
                    })
        except Exception as e:
            print(f"Preferences not found for Employee ID {employee_id} — skipping. Error: {e}")
            parsed_preferences = []

        try:
            skills_res = client.verint_call(f"wfo/user-mgmt-api/v1/employees/{employee_id}/skills")
            employee_skills = parse_employee_skills(skills_res)
            total_skills = len(skills_res.get("data", []))
            kept_skills = len(employee_skills) if employee_skills else 0
            print(f"Employee {employee_id}: kept {kept_skills} of {total_skills} skills after end‑date filter")
        except Exception as e:
            print(f"Skill fetch failed for Employee ID {employee_id} — skipping. Error: {e}")
            employee_skills = ""

        try:
            udf_res = client.verint_call(f"wfo/user-mgmt-api/v1/employees/{employee_id}/user-defined-fields")
            employee_udfs = parse_employee_udfs(udf_res)
        except Exception as e:
            print(f"UDF fetch failed for Employee ID {employee_id} — skipping. Error: {e}")
            employee_udfs = ""

        # Build compact JSON for the employee's *active* skills
        if employee_skills:
            skills_json = json.dumps(employee_skills, separators=(",", ":"))
        else:
            skills_json = None

        # Supervisor and Team Lead as JSON objects with id and name, use None for empty
        supervisor_obj = None
        try:
            supervisor_res = client.verint_call(f"wfo/user-mgmt-api/v1/employees/{employee_id}/supervisor")
            sup_data = supervisor_res.get("data", {})
            sup_attr = sup_data.get("attributes", {})
            sup_name = f"{sup_attr.get('firstName', '').strip()} {sup_attr.get('lastName', '').strip()}".strip()
            supervisor_obj = json.dumps({"id": sup_data.get("id"), "name": sup_name}) if sup_data.get("id") else None
        except Exception as e:
            print(f"Supervisor fetch failed for Employee ID {employee_id} — skipping. Error: {e}")

        teamlead_obj = None
        try:
            teamlead_res = client.verint_call(f"wfo/user-mgmt-api/v1/employees/{employee_id}/teamLead")
            tl_data = teamlead_res.get("data", {})
            tl_attr = tl_data.get("attributes", {})
            tl_name = f"{tl_attr.get('firstName', '').strip()} {tl_attr.get('lastName', '').strip()}".strip()
            teamlead_obj = json.dumps({"id": tl_data.get("id"), "name": tl_name}) if tl_data.get("id") else None
        except Exception as e:
            print(f"Team lead fetch failed for Employee ID {employee_id} — skipping. Error: {e}")

        record = {
            "Employee ID": employee_id,
            "Username": user.get("username"),
            "User Status": user.get("status"),
            "Employee Number": attr.get("employeeNumber"),
            "Employee Type": attr.get("employeeType"),
            "Job Title": job_title_json,
            "Is Supervisor": attr.get("isSupervisor"),
            "Is Team Lead": attr.get("isTeamLead"),
            "Organization ID": organization_id,
            "Organization Name": org_meta.get("name"),
            "First Name": person.get("firstName"),
            "Middle Initial": person.get("middleInitial"),
            "Last Name": person.get("lastName"),
            "Email": contact.get("email"),
            "Desktop Messaging Username": contact.get("desktopMessagingUsername"),
            "Supervisor": supervisor_obj,
            "Team Lead": teamlead_obj,
            "Home Phone": contact.get("homePhone"),
            "Work Phone": contact.get("workPhone"),
            "Cell Phone": contact.get("cellPhone"),
            "Start Time": attr.get("startTime"),
            "End Time": attr.get("endTime"),
            "SSN": person.get("ssn"),
            "Birth Date": person.get("birthDate"),
            "Address Line 1": address.get("addressLine1"),
            "Address Line 2": address.get("addressLine2"),
            "Address Line 3": address.get("addressLine3"),
            "City": address.get("city"),
            "State": address.get("stateName"),
            "Zip Code": address.get("zipCode"),
            "Country": address.get("country"),
            "Workspace Logins (dataSourceName - loginName)": ", ".join(workspace_logins),
            "Preferences": json.dumps(parsed_preferences) if parsed_preferences else None,
            "User Defined Fields": json.dumps(employee_udfs) if employee_udfs != "" else None,
            "Skills": skills_json,
            "Groups": json.dumps(employee_groups_map.get(employee_id, [])) or None
        }

        records.append(record)

    df = pd.DataFrame(records)

    # Write to shared workbook "output/verint_full_export.xlsx"
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb_path = "output/verint_full_export.xlsx"
    os.makedirs("output", exist_ok=True)

    if os.path.exists(wb_path):
        wb = load_workbook(wb_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # Always overwrite the previous “Employees” sheet if it exists
    if "Employees" in wb.sheetnames:
        del wb["Employees"]
    ws = wb.create_sheet(title="Employees")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    wb.save(wb_path)
    print(f"Employee data sheet written to {wb_path}")