"""
Module: group_extractor.py
Purpose:
    Extracts hierarchical group data from the Verint API, along with group
    membership, and exports it into an Excel sheet. Also builds a lookup of employee-to-group
    assignments for downstream processing.
"""

import pandas as pd
from verint_client import VerintClient
from datetime import datetime
import json
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import load_workbook

employee_groups_map = defaultdict(list)

def extract_groups():
    """
    Connects to Verint API to fetch and export group hierarchy and metadata.
    """
    client = VerintClient()
    response = client.verint_call("wfo/user-mgmt-api/v1/groups")

    # Call groups API and save response to disk for traceability
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs("json_dump", exist_ok=True)
    with open(f"json_dump/group_response_{timestamp}.json", "w") as f:
        json.dump(response, f, indent=4)

    groups = response.get("data", [])

    # Build maps for quick lookup and hierarchy traversal
    group_by_id = {group["id"]: group for group in groups}  
    children_map = defaultdict(list)  
    for group in groups:
        parent_id = group["attributes"].get("parentId")
        if parent_id is not None:
            children_map[str(parent_id)].append(group["id"])

    rows_detailed = []

    def get_group_members(group_id):
        """
        Retrieves the list of employees in a group and populates the global
        employee_groups_map with group assignments.
        """
        try:
            # API for group members
            members_response = client.verint_call(f"wfo/user-mgmt-api/v1/groups/{group_id}/employees")
            members_data = members_response.get("data", [])
            group_name = group_by_id[group_id]["attributes"].get("name", "")

            for emp in members_data:
                employee_groups_map[emp["id"]].append({
                    "id": group_id,
                    "name": group_name
                })
            # Member details for export
            return [
                {
                    "id": emp.get("id", ""),
                    "firstName": emp["attributes"].get("firstName", ""),
                    "lastName": emp["attributes"].get("lastName", ""),
                    "middleInitial": emp["attributes"].get("middleInitial", "")
                }
                for emp in members_data
            ]
        except Exception as e:
            print(f"Failed to fetch members for Group ID {group_id}: {e}")
            return []

    def fill_levels_detailed(group_id, level, path):
        """
        Recursively traverses the group hierarchy and collects relevant metadata
        and membership details for export.
        """
        group = group_by_id[group_id]
        name = group["attributes"].get("name", "")
        description = group["attributes"].get("description", "") or ""
        group_type = ", ".join(group["attributes"].get("groupType", []))

        # Get all members in this group
        members = get_group_members(group_id)

        # Initialize row with name placed at the correct level column
        entry = [''] * 10 
        group_id_str = str(group_id)
        entry[level] = name
        entry += [group_id_str, description, group_type, json.dumps(members)]
        rows_detailed.append(entry)

        # Recursively process child groups, increasing the hierarchy level
        for child_id in children_map.get(group_id, []):
            fill_levels_detailed(child_id, level + 1, path + [name])

    # Start recursion from root-level groups (those with no parent)
    for group in groups:
        if group["attributes"].get("parentId") is None:
            fill_levels_detailed(group["id"], 0, [])

    # Prepare Excel workbook and write Group Hierarchy sheet directly
    wb_path = "output/verint_full_export.xlsx"
    os.makedirs("output", exist_ok=True)

    if os.path.exists(wb_path):
        wb = load_workbook(wb_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "Group Hierarchy" in wb.sheetnames:
        std = wb["Group Hierarchy"]
        wb.remove(std)
    ws = wb.create_sheet(title="Group Hierarchy")

    headers = [f"Level {i+1}" for i in range(10)] + ["Group ID", "Description", "Group Type", "Group Members"]
    ws.append(headers)
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    for row in rows_detailed:
        ws.append(row)

    wb.save(wb_path)
    print(f"Group hierarchy sheet written to {wb_path}")
    return employee_groups_map

if __name__ == "__main__":
    extract_groups()
