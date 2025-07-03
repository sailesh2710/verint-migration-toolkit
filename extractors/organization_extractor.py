import os
import json
import pandas as pd
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from verint_client import VerintClient


def extract_organizations():
    client = VerintClient()

    # Call organizations API
    org_response = client.verint_call("wfo/user-mgmt-api/v1/organizations")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs("json_dump", exist_ok=True)
    with open(f"json_dump/org_response_{timestamp}.json", "w") as f:
        json.dump(org_response, f, indent=4)

    orgs = org_response.get("data", [])
    org_by_id = {org["id"]: org for org in orgs}
    children_map = defaultdict(list)
    for org in orgs:
        parent_id = org["attributes"].get("parentId")
        if parent_id is not None:
            children_map[str(parent_id)].append(org["id"])

    rows_hierarchy = []

    def add_org_and_skills(org_id, level):
        org = org_by_id[org_id]
        attr = org["attributes"]

        row = [''] * 10
        org_id_str = str(org_id)
        row[level] = attr.get("name", "")

        description = attr.get("description", "")
        timezone = attr.get("timeZone", "")
        week_start = attr.get("weekStartDay", "")
        seats = attr.get("seatsNumber", "")
        location = attr.get("location", "")

        # Get skills assigned directly to this org
        skill_response = client.verint_call(f"wfo/user-mgmt-api/v1/organizations/{org_id}/skills")
        skills = skill_response.get("data", [])
        direct_skills = []

        for skill in skills:
            org_ref_id = skill.get("relationships", {}).get("organization", {}).get("data", {}).get("id")
            if str(org_ref_id) == str(org_id):
                sattr = skill.get("attributes", {})
                direct_skills.append({
                    "name": sattr.get("name", ""),
                    "media": sattr.get("media", ""),
                    "description": sattr.get("description", ""),
                    "isActive": sattr.get("isActive", False)
                })

        # Get UDFs assigned directly to this org
        try:
            udf_response = client.verint_call(f"wfo/user-mgmt-api/v1/organizations/{org_id}/user-defined-fields")
            udfs = udf_response.get("data", [])
            udf_list = []
            for udf in udfs:
                org_ref_id = udf.get("relationships", {}).get("organization", {}).get("data", {}).get("id")
                if str(org_ref_id) == str(org_id):
                    uattr = udf.get("attributes", {})
                    udf_list.append({
                        "name": uattr.get("name", ""),
                        "description": uattr.get("description", ""),
                        "udfType": uattr.get("udfType", ""),
                        "values": uattr.get("values", []) if "values" in uattr else []
                    })
        except Exception as e:
            print(f"UDF fetch failed for Org ID {org_id} — skipping. Error: {e}")
            udf_list = []

        # Get Job Titles assigned directly to this org
        try:
            job_response = client.verint_call(f"wfo/user-mgmt-api/v1/organizations/{org_id}/jobTitles")
            jobs = job_response.get("data", [])
            job_list = []
            for job in jobs:
                org_ref_id = job.get("relationships", {}).get("organization", {}).get("data", {}).get("id")
                if str(org_ref_id) == str(org_id):
                    jattr = job.get("attributes", {})
                    job_list.append({
                        "name": jattr.get("name", ""),
                        "description": jattr.get("description", "")
                    })
        except Exception as e:
            print(f"Job Title fetch failed for Org ID {org_id} — skipping. Error: {e}")
            job_list = []

        row += [description, timezone, week_start, seats, location,
                json.dumps(direct_skills), json.dumps(udf_list), json.dumps(job_list)]
        rows_hierarchy.append(row[:10] + [org_id_str] + row[10:])

        for child_id in children_map.get(org_id, []):
            add_org_and_skills(child_id, level + 1)

    for org in orgs:
        if org["attributes"].get("parentId") is None:
            add_org_and_skills(org["id"], 0)

    # Create workbook and write Organization Hierarchy sheet directly
    wb_path = "output/verint_full_export.xlsx"
    os.makedirs("output", exist_ok=True)

    if os.path.exists(wb_path):
        from openpyxl import load_workbook
        wb = load_workbook(wb_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "Organization Hierarchy" in wb.sheetnames:
        std = wb["Organization Hierarchy"]
        wb.remove(std)
    ws = wb.create_sheet(title="Organization Hierarchy")   

    headers = [f"Level {i+1}" for i in range(10)] + ["Organization ID"] + [
        "Description", "TimeZone", "WeekStartDay", "SeatsNumber", "Location",
        "Skills (Direct Only)", "User Defined Fields (Direct Only)", "Job Titles (Direct Only)"
    ]
    ws.append(headers)
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.fill = PatternFill(start_color="FBE4D5", end_color="FBE4D5", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    for row in rows_hierarchy:
        ws.append(row)

    wb.save(wb_path)
    print(f"Organization hierarchy sheet written to {wb_path}")

if __name__ == "__main__":
    extract_organizations()