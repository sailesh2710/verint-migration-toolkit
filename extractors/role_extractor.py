"""
Module: role_extractor.py
Purpose:
    Extracts role metadata from the Verint API and exports it into an Excel sheet.
    Captures attributes such as role name, description, admin/default flags, and owning organization.
"""

import os
import json
import pandas as pd
from datetime import datetime
from verint_client import VerintClient

def extract_roles():
    client = VerintClient()

    # Call roles API endpoint to retrieve all roles
    response = client.verint_call("wfo/user-mgmt-api/v1/roles")

    # Save full JSON response to disk with timestamp for traceability
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs("json_dump", exist_ok=True)
    with open(f"json_dump/roles_response_{timestamp}.json", "w") as f:
        json.dump(response, f, indent=4)

    roles = response.get("data", [])
    records = []

    # Extract relevant metadata from each role entry
    for role in roles:
        attr = role.get("attributes", {})
        rel = role.get("relationships", {}).get("organization", {}).get("data", {})
        meta = rel.get("meta", {})

        records.append({
            "Role Name": attr.get("name"),
            "Description": attr.get("description", ""),
            "Is Default": attr.get("isDefault", False),
            "Is Admin Role": attr.get("isAdminRole", False),
            "Organization Name": meta.get("name")
        })

    df = pd.DataFrame(records)

    # Load or create the Excel workbook
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb_path = "output/verint_full_export.xlsx"
    os.makedirs("output", exist_ok=True)

    if os.path.exists(wb_path):
        wb = load_workbook(wb_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "Roles" in wb.sheetnames:
        std = wb["Roles"]
        wb.remove(std)
    ws = wb.create_sheet(title="Roles")
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    wb.save(wb_path)
    print(f"Roles sheet written to {wb_path}")

if __name__ == "__main__":
    extract_roles()
