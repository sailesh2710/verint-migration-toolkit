"""
Module: access_rights_extractor.py
Purpose:
    Extracts access rights data for each employee from the Verint API,
    including assigned roles, admin flags, default status, owning organization,
    and lists of accessible organizations and groups.
    Outputs the data into an Excel sheet named 'Access Rights'.
"""

import os
import json
import pandas as pd
from datetime import datetime
from verint_client import VerintClient

def extract_access_rights():
    client = VerintClient()

    # Fetch all employees from Verint
    emp_response = client.verint_call("wfo/user-mgmt-api/v1/employees")
    employees = emp_response.get("data", [])

    records = []  # Will store formatted access rights data for all employees
    
    for emp in employees:
        emp_id = emp.get("id")
        emp_name = emp.get("attributes", {}).get("user", {}).get("username", "")

        try:
            # Fetch all roles assigned to this employee
            response = client.verint_call(f"wfo/user-mgmt-api/v1/employees/{emp_id}/roles")
            roles = response.get("data", [])

            for role in roles:
                attr = role.get("attributes", {})
                rel = role.get("relationships", {})

                # Get organization that owns the role
                org_creator = rel.get("organization", {}).get("data", {})
                org_creator_name = org_creator.get("meta", {}).get("name", "")

                # Gather accessible organizations under this role
                accessible_orgs = [
                    {
                        "id": o.get("id"),
                        "name": o.get("meta", {}).get("name")
                    }
                    for o in rel.get("organizations", {}).get("data", [])
                ]

                # Gather accessible groups under this role
                accessible_groups = [
                    {
                        "id": g.get("id"),
                        "name": g.get("meta", {}).get("name")
                    }
                    for g in rel.get("groups", {}).get("data", [])
                ]

                # Append consolidated access info for this role
                records.append({
                    "Employee ID": emp_id,
                    "Username": emp_name,
                    "Role Name": attr.get("name"),
                    "Description": attr.get("description", ""),
                    "Is Admin Role": attr.get("isAdminRole", False),
                    "Is Default": attr.get("isDefault", False),
                    "Owning Org Name": org_creator_name,
                    "Accessible Orgs": json.dumps(accessible_orgs),
                    "Accessible Groups": json.dumps(accessible_groups)
                })

        except Exception as e:
            print(f"Access rights not found for Employee ID {emp_id} — skipping. Error: {e}")

    # Write results into an Excel workbook under "Access Rights" sheet
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    df = pd.DataFrame(records)
    wb_path = "output/verint_full_export.xlsx"
    os.makedirs("output", exist_ok=True)

    if os.path.exists(wb_path):
        wb = load_workbook(wb_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "Access Rights" in wb.sheetnames:
        std = wb["Access Rights"]
        wb.remove(std)
    ws = wb.create_sheet(title="Access Rights")

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    wb.save(wb_path)
    print(f"Access rights sheet written to {wb_path}")

if __name__ == "__main__":
    extract_access_rights()